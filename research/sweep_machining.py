#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sweep_machining.py — 扫描 (容差, 分片数) → 加工时间数据，输出 xlsx，支持断点重启。

工作流：对每个 (tolerance, nSplitU, nSplitV) 组合：
  1) 调 simple.exe 做直纹面分片拟合（默认固定分片 --no-refine，指定分片数量）
  2) 读 *_params.txt + meta.json，用 compute_machining 计算侧铣/点铣加工时间
  3) 记录 (容差, 分片数, 拟合误差, 扭转角, 侧铣时间, 点铣时间, 提速比)
结果实时写入 checkpoint JSON（断点重启用）和 xlsx。

用法:
  python sweep_machining.py <file1> <file2> \
      --out result.xlsx --checkpoint sweep_ckpt.json \
      --tolerances 0.05,0.1,0.2,0.5 \
      --splits "1,1;2,2;3,3;4,4;6,6;8,8;10,10" \
      [--no-refine] [--workdir tmp]
"""
import os
import sys
import json
import argparse
import subprocess
import shutil
import tempfile
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
import compute_machining as cm

PROJECT_DIR = Path(__file__).resolve().parent.parent
BUILD_EXE = PROJECT_DIR / "build" / "Release" / "simple.exe"

HEADERS = [
    "total", "num_combos", "num_patches", "flank_regions",
    "max_error_mm", "mean_error_mm", "q95_error_mm", "rms_error_mm",
    "max_twist_deg",
    "flank_cut_s", "flank_overhead_s", "flank_total_s",
    "point_cut_s", "point_total_s", "speedup",
]


def _key(total):
    return str(total)


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def run_fitting(file1, file2, outdir, tol, nu, nv, ps_ranges=None, face_idx=0):
    cmd = [str(BUILD_EXE), file1, file2, "--mode", "ruled",
           "--outdir", outdir, "--tolerance", str(tol),
           "--nsplit-u", str(nu), "--nsplit-v", str(nv),
           "--no-refine", "--balance-edges", "--max-cells", "200000"]
    if ps_ranges:
        for k, (dir_, ranges) in enumerate(ps_ranges, start=1):
            if not ranges:
                continue
            rk = 'v' if dir_ == 'V' else 'u'
            us, ue = ranges[0]   # 取第一个叶盆/叶背区间
            cmd += [f"--face-idx{k}", str(face_idx), f"--{rk}-range{k}", f"{us},{ue}"]
    r = subprocess.run(cmd, cwd=str(PROJECT_DIR), capture_output=True,
                       text=True, encoding="utf-8", errors="replace", timeout=900)
    if r.returncode != 0:
        err = (r.stderr or "").strip().splitlines()
        if err:
            log(f"  [fitting stderr] {'; '.join(err[-3:])}")
    return r.returncode


def run_exe_json(args_list):
    """运行 simple.exe 并从 stdout 解析 JSON（跳过前面的日志行）。"""
    r = subprocess.run([str(BUILD_EXE)] + args_list, cwd=str(PROJECT_DIR),
                       capture_output=True, text=True, encoding="utf-8",
                       errors="replace", timeout=300)
    s = r.stdout.strip()
    i = s.find('{')
    if i >= 0:
        s = s[i:]
    try:
        return json.loads(s)
    except Exception:
        return None


def auto_identify(filepath):
    """auto-identify 找叶片面，返回 (pressureIndex, suctionIndex)。"""
    info = run_exe_json([filepath, "--mode", "auto-identify"])
    if not info or not info.get('success'):
        return 0, 0
    return info.get('pressureIndex', 0), info.get('suctionIndex', 0)


def split_blade(filepath, face_idx):
    """split-blade 一个面，返回 (dir, [regions])。"""
    info = run_exe_json([filepath, "--mode", "split-blade",
                         "--face-idx", str(face_idx)])
    if not info or not info.get('success'):
        return None, []
    return info.get('dir', 'V'), info.get('regions', [])


def identify_ps(regions):
    """从 split-blade 结果里挑出叶盆/叶背区间，返回 (pressure, suction) 两个 (u_start, u_end)。"""
    pressure = [r for r in regions if r.get('label') == 'pressure']
    suction = [r for r in regions if r.get('label') == 'suction']
    nonedge = [r for r in regions if r.get('label') != 'edge']
    if not pressure and not suction:
        # 标签不全时，按宽度取两个非 edge 区间
        nonedge.sort(key=lambda r: r.get('uEnd', 0) - r.get('uStart', 0), reverse=True)
        if len(nonedge) >= 2:
            return (nonedge[0]['uStart'], nonedge[0]['uEnd']), \
                   (nonedge[1]['uStart'], nonedge[1]['uEnd'])
        if nonedge:
            return (nonedge[0]['uStart'], nonedge[0]['uEnd']), None
        return None, None
    p = (pressure[0]['uStart'], pressure[0]['uEnd']) if pressure else None
    s = (suction[0]['uStart'], suction[0]['uEnd']) if suction else None
    return p, s


def read_error_stats(outdir):
    """从 meta.json 读取误差分布统计（max/mean/Q95/rms，按每格 maxErr 计算）。"""
    meta = os.path.join(outdir, "meta.json")
    if not os.path.exists(meta):
        return {}
    try:
        with open(meta, encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return {}
    errs = []
    for s in d.get("surfaces", []):
        errs.extend(c.get("maxErr", 0.0) for c in s.get("cells", []))
    if not errs:
        return {}
    errs.sort()
    n = len(errs)
    return {
        "max_error_mm": round(errs[-1], 6),
        "mean_error_mm": round(sum(errs) / n, 6),
        "q95_error_mm": round(errs[min(n - 1, int(n * 0.95))], 6),
        "rms_error_mm": round((sum(e * e for e in errs) / n) ** 0.5, 6),
    }


def collect(outdir, args):
    patches = cm.compute(outdir, args)
    if not patches:
        return None
    s = cm.summarize(patches, args)
    max_twist = max((p.twist for p in patches), default=0.0)
    speedup = None
    if s["flank"]["total"] > 0:
        speedup = round(s["point"]["total"] / s["flank"]["total"], 3)
    return {
        "num_patches": s["num_patches"],
        "flank_regions": s["flank_regions"],
        "max_twist_deg": round(max_twist, 3),
        "flank_cut_s": round(s["flank"]["cut"], 2),
        "flank_overhead_s": round(s["flank"]["overhead"], 2),
        "flank_total_s": round(s["flank"]["total"], 2),
        "point_cut_s": round(s["point"]["cut"], 2),
        "point_total_s": round(s["point"]["total"], 2),
        "speedup": speedup,
    } | read_error_stats(outdir)


def write_xlsx(rows, out_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "machining_sweep"
    ws.append(HEADERS)
    for r in rows:
        ws.append([r.get(h) for h in HEADERS])
    wb.save(out_path)


def add_charts(out_path):
    """在 xlsx 里加图表：提速 vs 总分片数、误差 vs 总分片数。"""
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    try:
        wb = openpyxl.load_workbook(out_path)
    except Exception:
        return
    ws = wb.active
    if ws.max_row < 2:
        return

    if "charts" in wb.sheetnames:
        del wb["charts"]

    headers = [c.value for c in ws[1]]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    try:
        x_idx = headers.index("total")
        speedup_idx = headers.index("speedup")
        err_cols = [headers.index(x) for x in
                    ("max_error_mm", "mean_error_mm", "q95_error_mm", "rms_error_mm")]
    except ValueError:
        return
    rows.sort(key=lambda r: r[x_idx] if r[x_idx] is not None else 0)

    cs = wb.create_sheet("charts")
    cs.append(headers)
    for r in rows:
        cs.append(r)

    # 图1：提速 vs 总分片数（纯折线）
    c1 = LineChart()
    c1.title = "Speedup vs Total Patches"
    c1.y_axis.title = "Speedup"
    c1.x_axis.title = "Total Patches"
    c1.width = 22
    c1.height = 11
    c1.x_axis.tickLblSkip = 5
    c1.x_axis.tickMarkSkip = 5
    ref_x = Reference(cs, min_col=x_idx + 1, min_row=2, max_row=cs.max_row)
    ref_y = Reference(cs, min_col=speedup_idx + 1, min_row=1, max_row=cs.max_row)
    c1.add_data(ref_y, titles_from_data=True)
    c1.set_categories(ref_x)
    cs.add_chart(c1, "A" + str(cs.max_row + 3))

    # 图2：误差 vs 总分片数
    c2 = LineChart()
    c2.title = "Error vs Total Patches"
    c2.y_axis.title = "Error (mm)"
    c2.x_axis.title = "Total Patches"
    c2.width = 22
    c2.height = 11
    c2.x_axis.tickLblSkip = 5
    c2.x_axis.tickMarkSkip = 5
    ref_x2 = Reference(cs, min_col=x_idx + 1, min_row=2, max_row=cs.max_row)
    for ci in err_cols:
        ref = Reference(cs, min_col=ci + 1, min_row=1, max_row=cs.max_row)
        c2.add_data(ref, titles_from_data=True)
    c2.set_categories(ref_x2)
    cs.add_chart(c2, "A" + str(cs.max_row + 22))

    wb.save(out_path)


def load_xlsx_rows(path):
    """读取已有 xlsx 的行（用于断点追加，不丢失历史数据）。"""
    import openpyxl
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = None
        rows = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = list(row)
                continue
            d = dict(zip(headers, row))
            if d.get("total") is not None:
                rows.append(d)
        return rows
    except Exception:
        return []


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass
    ap = argparse.ArgumentParser(description="单文件叶片：auto-identify→split-blade→拟合扫描(总分片数范围)，输出 xlsx")
    ap.add_argument("file", help="叶片文件（单文件，如 blade.igs）")
    ap.add_argument("--out", required=True, help="输出 xlsx 路径")
    ap.add_argument("--checkpoint", default=None, help="断点检查点 JSON 路径")
    ap.add_argument("--tolerance", type=float, default=0.1, help="拟合容差 mm")
    ap.add_argument("--total-min", type=int, default=4, help="总分片数下限（含）")
    ap.add_argument("--total-max", type=int, default=100, help="总分片数上限（含）")
    ap.add_argument("--workdir", default=None, help="临时工作目录（默认系统临时目录）")
    args = ap.parse_args()

    if not BUILD_EXE.exists():
        log(f"[ERROR] simple.exe not found: {BUILD_EXE}")
        sys.exit(1)

    # 模拟 UI 手动流程：auto-identify → split-blade → identify pressure/suction，缓存结果
    log("auto-identify...")
    pidx, sidx = auto_identify(args.file)
    log(f"  pressureIndex={pidx} suctionIndex={sidx}")

    split_face_idx = pidx if pidx >= 0 else 0
    log(f"split-blade face {split_face_idx}...")
    split_dir, regions = split_blade(args.file, split_face_idx)
    log(f"  dir={split_dir} regions={len(regions)}")
    for reg in regions:
        log(f"    {reg.get('label')} V[{reg.get('uStart')},{reg.get('uEnd')}]")

    pressure, suction = identify_ps(regions)
    log(f"  pressure={pressure} suction={suction}")

    ps_ranges = None
    if pressure or suction:
        ps_ranges = [
            (split_dir, [pressure] if pressure else []),
            (split_dir, [suction] if suction else []),
        ]
    else:
        log("  未识别到叶盆/叶背区间，将拟合整面")

    mcfg = cm.load_config()
    mach_args = argparse.Namespace(
        feed=mcfg.get("feed", 500.0), tool_r=mcfg.get("tool_r", 5.0),
        ball_r=mcfg.get("ball_r", 5.0), scallop=mcfg.get("scallop", 0.1),
        twist_limit=mcfg.get("twist_limit", 2.0),
        taper_angle=mcfg.get("taper_angle", 3.0),
        overhead=mcfg.get("overhead", 4.0),
        point_overhead=mcfg.get("point_overhead", 10.0))

    # 断点重启：合并「已有 xlsx」+「checkpoint」，按 total 去重
    def key_of(r):
        if r.get("key"):
            return r["key"]
        return _key(r.get("total"))

    merged = {}
    for r in load_xlsx_rows(args.out):
        r.setdefault("key", key_of(r))
        merged[r["key"]] = r
    if args.checkpoint and os.path.exists(args.checkpoint):
        try:
            with open(args.checkpoint, encoding="utf-8") as f:
                ck = json.load(f).get("results", [])
            for r in ck:
                r.setdefault("key", key_of(r))
                merged[r["key"]] = r
        except Exception:
            pass
    results = list(merged.values())
    done_keys = {r["key"] for r in results if r.get("key")}

    workdir_base = args.workdir or tempfile.mkdtemp(prefix="sweep_")
    os.makedirs(workdir_base, exist_ok=True)

    totals = list(range(args.total_min, args.total_max + 1))
    log(f"扫描总分片数 {args.total_min}~{args.total_max}（{len(totals)} 个），"
        f"每个 total 枚举所有乘法组合并取平均")
    log(f"simple.exe: {BUILD_EXE}")

    def enumerate_splits(total):
        combos = []
        for nu in range(1, total + 1):
            if total % nu == 0:
                combos.append((nu, total // nu))
        return combos

    def save():
        if args.checkpoint:
            try:
                with open(args.checkpoint, "w", encoding="utf-8") as f:
                    json.dump({"results": results}, f, ensure_ascii=False, indent=1)
            except Exception as e:
                log(f"  [warn] 写 checkpoint 失败: {e}")
        try:
            write_xlsx(results, args.out)
        except Exception as e:
            log(f"  [warn] 写 xlsx 失败: {e}")

    def average_rows(rows):
        if not rows:
            return None
        keys = [k for k in rows[0].keys() if k != "key"]
        avg = {}
        for k in keys:
            vals = [r.get(k) for r in rows if r.get(k) is not None]
            if vals and all(isinstance(v, (int, float)) for v in vals):
                avg[k] = round(sum(vals) / len(vals), 4)
            else:
                avg[k] = rows[0].get(k)
        return avg

    i = 0
    for total in totals:
        key = _key(total)
        i += 1
        if key in done_keys:
            log(f"[{i}/{len(totals)}] skip (已完成): total={total}")
            continue
        combos = enumerate_splits(total)
        log(f"[{i}/{len(totals)}] total={total} → {len(combos)} 组合 {combos}")
        combo_rows = []
        for (nu, nv) in combos:
            outdir = os.path.join(workdir_base, f"t{total}_u{nu}_v{nv}")
            if os.path.isdir(outdir):
                shutil.rmtree(outdir, ignore_errors=True)
            os.makedirs(outdir, exist_ok=True)
            try:
                rc = run_fitting(args.file, args.file, outdir, args.tolerance, nu, nv,
                                 ps_ranges, split_face_idx)
            except Exception as e:
                log(f"    [{nu}x{nv}] 拟合执行异常: {e}")
                continue
            if rc != 0:
                log(f"    [{nu}x{nv}] 拟合失败 rc={rc}")
                continue
            rec = collect(outdir, mach_args)
            if rec is None:
                log(f"    [{nu}x{nv}] 未解析到准线文件")
                continue
            combo_rows.append(rec)
        if not combo_rows:
            log(f"  total={total} 所有组合失败，跳过")
            continue
        avg = average_rows(combo_rows)
        avg["key"] = key
        avg["total"] = total
        avg["num_combos"] = len(combo_rows)
        results.append(avg)
        log(f"  → avg: patches={avg['num_patches']} maxErr={avg['max_error_mm']}mm "
            f"flank={avg['flank_total_s']}s point={avg['point_total_s']}s 提速={avg['speedup']}x")
        save()

    add_charts(args.out)
    log(f"完成。共 {len(results)} 条结果 → {args.out}")
    if args.checkpoint:
        log(f"断点文件 → {args.checkpoint}")


if __name__ == "__main__":
    main()
